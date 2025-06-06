import pandas as pd
import numpy as np
from employees import employee
from timeconverter import convertTo24, convertTo12
from openpyxl import load_workbook
from datetime import datetime

def get_employeeDict():

        employeeDict = {}

        scheduleRaw = pd.read_excel("data/example.xlsx", header = None, skiprows=9)
        schedule = scheduleRaw.copy()
        #print(schedule.to_string())

        schedule[8] = schedule[8].replace(np.nan, '0')
        #/Users/derek/Documents/vova/zone chart/zonechart/data/employee test.xlsx

        for  i in range(len(schedule)):
            if schedule.iloc[i, 8] != '0' and schedule.iloc[i, 8] != "Employee" and schedule.iloc[i, 8] not in employeeDict:
                time = schedule.iloc[i, 10]
                timeSplit = time.split("-")
                start_time = convertTo24(timeSplit[0].strip())
                end_time = convertTo24(timeSplit[1].strip())
                employeeDict[schedule.iloc[i, 8]] = [start_time, end_time]
            elif schedule.iloc[i, 8] in employeeDict:
                time = schedule.iloc[i, 10]
                timeSplit = time.split("-")
                end_time = convertTo24(timeSplit[1].strip())
                old_time = employeeDict[schedule.iloc[i, 8]]
                start_time = old_time[0].strip()
                employeeDict[schedule.iloc[i, 8]] = [start_time, end_time]
            else:
                continue

        
        return employeeDict

def dictToSorted2DArray(employeeDict):
    employeeArray = [[0, 0, 0] for _ in range(len(employeeDict))]
    for i, (key, value) in enumerate(employeeDict.items()):
        employeeArray[i][0] = key
        employeeArray[i][1] = value[0]  # Start time
        employeeArray[i][2] = value[1]  # End time

    employeeArray.sort(key=lambda x: x[1])  # Sort by start time
    print(employeeArray)

    return employeeArray

    
            

def writeToExcel(day):

    file_path = "data/playground copy.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Monday"]  

    employeeArray = dictToSorted2DArray(get_employeeDict())

    excelCount = 18

    shift = 0
    if day == "weekday":
        shift = 0
    elif day == "saturday":
        shift = 1
    elif day == "sunday":
        shift = 2

    for i in range(len(employeeArray)): 
        values = employeeArray[i]
        ws[f"H{excelCount}"] = values[0]
        ws[f"I{excelCount}"] = f"{convertTo12(values[1])} - {convertTo12(values[2])}"

        
        
        if i > 0:
            shift += int(employeeArray[i][1].split(":")[0]) - int(employeeArray[i-1][1].split(":")[0])
        timeEnd = int(employeeArray[i][2].split(":")[0])
        timeEndMinutes = int(employeeArray[i][2].split(":")[1])
        if timeEndMinutes > 30:
            timeEnd += 1
        timeStart = int(employeeArray[i][1].split(":")[0])
        delta =  timeEnd - timeStart
        for j in range(delta):
            char = chr(74 + j +  shift)  # J is 74 in ASCII
            ws[f"{char}{excelCount}"] = "1"
            print(delta, char, shift)
        print(delta, char, shift, employeeArray[i][0], employeeArray[i][1], employeeArray[i][2])
            
        excelCount += 1
    # Save the workbook

    

    wb.save(file_path)
    


writeToExcel("weekday")

