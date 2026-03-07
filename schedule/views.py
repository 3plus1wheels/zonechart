import tempfile
import os
import traceback as tb

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
import openpyxl

from .models import Employee, Shift, StaffZone, ZONE_FIELDS
from .serializers import ShiftSerializer, EmployeeSerializer, StaffZoneSerializer
from .parsers import parse_schedule_xlsx

WORKBOOK_HOURS = list(range(8, 21))  # 8 am … 8-9 pm slot

WORKBOOK_COL_HEADERS = [
    '8am-9am','9am-10am','10am-11am','11am-12pm',
    '12pm-1pm','1pm-2pm','2pm-3pm','3pm-4pm',
    '4pm-5pm','5pm-6pm','6pm-7pm','7pm-8pm','8pm-9pm',
]

# Zone assignment order for Stylists each hour.
# Slots are filled left-to-right; each Stylist is used at most once per hour.
SLOT_SEQUENCE = ['womens', 'mens', 'fits', 'cash', 'fits', 'mens', 'womens', 'greet', 'mens', 'womens']

# Main zones that managers can plug when stylists don't fill them
MAIN_ZONE_SLOTS = ['womens', 'mens', 'fits', 'cash']


# ---------------------------------------------------------------------------
# Helper: save uploaded file to a temp path
# ---------------------------------------------------------------------------
def _monday_of(records: list) -> str:
    """Return the ISO Monday date of the earliest shift date in records."""
    from datetime import timedelta
    dates = [r['date'] for r in records if r.get('date')]
    if not dates:
        return ''
    earliest = min(dates)
    # weekday(): Mon=0 … Sun=6
    monday = earliest - timedelta(days=earliest.weekday())
    return monday.isoformat()


def _save_temp(xlsx_file):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        for chunk in xlsx_file.chunks():
            tmp.write(chunk)
        return tmp.name


class ShiftListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_filter = request.query_params.get('date')
        week_start = request.query_params.get('week_start')

        shifts = Shift.objects.select_related('employee').all()

        if date_filter:
            shifts = shifts.filter(date=date_filter)
        elif week_start:
            from datetime import datetime, timedelta
            try:
                start = datetime.strptime(week_start, '%Y-%m-%d').date()
                end = start + timedelta(days=6)
                shifts = shifts.filter(date__range=[start, end])
            except ValueError:
                pass

        serializer = ShiftSerializer(shifts, many=True)
        return Response(serializer.data)


class EmployeeListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        employees = Employee.objects.prefetch_related('shifts').all()
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)


class ImportScheduleView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        xlsx_file = request.FILES.get('file')
        if not xlsx_file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        if not xlsx_file.name.endswith('.xlsx'):
            return Response({'error': 'File must be an .xlsx file.'}, status=status.HTTP_400_BAD_REQUEST)

        clear = request.data.get('clear', 'false').lower() == 'true'

        tmp_path = _save_temp(xlsx_file)

        try:
            records = parse_schedule_xlsx(tmp_path)
        except Exception as exc:
            trace = tb.format_exc()
            return Response(
                {'error': f'Failed to parse XLSX: {exc}', 'traceback': trace},
                status=status.HTTP_400_BAD_REQUEST,
            )
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

        parsed_count = len(records)
        if not records:
            return Response({
                'message': 'No shifts found in file. The XLSX layout may not match the expected format.',
                'shifts_created': 0,
                'shifts_updated': 0,
                'employees_created': 0,
                'parsed_count': 0,
            })

        employees_created = 0
        shifts_created = 0
        shifts_updated = 0

        with transaction.atomic():
            if clear:
                Shift.objects.all().delete()

            for record in records:
                if not record.get('date'):
                    continue

                employee, emp_new = Employee.objects.update_or_create(
                    name=record['employee_name'],
                    defaults={'primary_job': record['primary_job']},
                )
                if emp_new:
                    employees_created += 1

                _, shift_new = Shift.objects.update_or_create(
                    employee=employee,
                    date=record['date'],
                    start_time=record['start_time'],
                    defaults={
                        'end_time': record['end_time'],
                        'day_label': record['day_label'],
                        'role': record['role'],
                    },
                )
                if shift_new:
                    shifts_created += 1
                else:
                    shifts_updated += 1

        return Response({
            'message': 'Import successful.',
            'employees_created': employees_created,
            'shifts_created': shifts_created,
            'shifts_updated': shifts_updated,
            'parsed_count': parsed_count,
            'week_start': _monday_of(records),
        }, status=status.HTTP_201_CREATED)


class WorkbookView(APIView):
    """GET /api/schedule/workbook/?week_start=YYYY-MM-DD&day=Mon
    Returns zone-chart rows sorted by start_time for one day."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import datetime, timedelta
        week_start = request.query_params.get('week_start')
        day = request.query_params.get('day', 'Mon')

        DAY_ORDER = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        if not week_start:
            return Response({'error': 'week_start required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            start = datetime.strptime(week_start, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid week_start'}, status=status.HTTP_400_BAD_REQUEST)

        day_offset = DAY_ORDER.index(day) if day in DAY_ORDER else 0
        target_date = start + timedelta(days=day_offset)

        shifts = (
            Shift.objects.select_related('employee')
            .filter(date=target_date)
            .order_by('start_time')
        )

        def _fmt(h, m):
            ah = h % 12 or 12
            return f"{ah}:{m:02d}" if m else str(ah)

        def _name_parts(raw):
            """Return (first, last_initial) from 'Last, First' format."""
            if ',' in raw:
                last, first = raw.split(',', 1)
                return first.strip().upper(), last.strip()[0].upper() if last.strip() else ''
            return raw.strip().upper(), ''

        # Determine which first names are shared so we can disambiguate
        shifts = list(shifts)  # evaluate queryset once
        # Sort: non-BOH by start_time first, BOH shifts always last
        shifts.sort(key=lambda s: (1 if s.role == 'BOH' else 0, s.start_time))
        from collections import Counter
        first_name_counts = Counter(_name_parts(s.employee.name)[0] for s in shifts)

        # --- Zone assignment for Stylists ---
        # For each hour, rank available Stylists by skill level in each slot
        # and greedily assign the best available person to each slot.
        emp_ids = [s.employee_id for s in shifts]
        staff_skill = {
            sz.employee_id: sz
            for sz in StaffZone.objects.filter(employee_id__in=emp_ids)
        }

        def _skill_sum(employee_id):
            sz = staff_skill.get(employee_id)
            if not sz:
                return 0
            return sum(getattr(sz, f, 0) or 0 for f in ZONE_FIELDS)

        hour_assignments = {}  # {h: {employee_id: zone_label_upper}}
        for h in WORKBOOK_HOURS:
            stylists_this_hour = [
                s for s in shifts
                if s.role == 'Stylist'
                and (s.start_time.hour * 60 + s.start_time.minute) < (h + 1) * 60
                and (s.end_time.hour   * 60 + s.end_time.minute)   > h * 60
            ]
            available = list(stylists_this_hour)
            assigned  = {}
            for slot_zone in SLOT_SEQUENCE:
                if not available:
                    break
                best = max(
                    available,
                    key=lambda s: (
                        getattr(staff_skill.get(s.employee_id), slot_zone, 0) or 0,
                        -_skill_sum(s.employee_id),   # tiebreaker: smaller total sum wins
                    )
                )
                assigned[best.employee_id] = slot_zone.upper()
                available.remove(best)
            for s in available:
                assigned[s.employee_id] = 'STYLIST'
            hour_assignments[h] = assigned

        # --- Manager (CEL) zone assignment ---
        # Process hours right-to-left.
        # Exactly ONE manager gets CEL per hour; the rest get FLEX.
        # Spread CEL evenly: prefer the manager with fewest consecutive CEL hours,
        # then fewest total CEL hours, then latest end_time as tiebreaker.
        cel_consecutive = {}   # {employee_id: int} resets when not working or assigned FLEX
        cel_total       = {}   # {employee_id: int} cumulative CEL hours
        all_mgr_ids = {s.employee_id for s in shifts if s.role == 'CEL'}

        # Store opens at 11am Sunday, 10am every other day
        store_open_hour = 11 if day == 'Sun' else 10

        for h in reversed(WORKBOOK_HOURS):
            managers_this_hour = [
                s for s in shifts
                if s.role == 'CEL'
                and (s.start_time.hour * 60 + s.start_time.minute) < (h + 1) * 60
                and (s.end_time.hour   * 60 + s.end_time.minute)   > h * 60
            ]

            # Reset consecutive count for managers not on shift this hour
            working_mgr_ids = {m.employee_id for m in managers_this_hour}
            for eid in all_mgr_ids:
                if eid not in working_mgr_ids:
                    cel_consecutive[eid] = 0

            available_mgrs = list(managers_this_hour)
            if not available_mgrs:
                continue

            # Before store opens → all managers on TASK
            if h < store_open_hour:
                for mgr in available_mgrs:
                    hour_assignments[h][mgr.employee_id] = 'TASK'
                continue

            for m in available_mgrs:
                cel_consecutive.setdefault(m.employee_id, 0)
                cel_total.setdefault(m.employee_id, 0)

            # Step 1: pick ONE manager for CEL (always, every hour)
            def _cel_priority(m):
                eid = m.employee_id
                consec = cel_consecutive.get(eid, 0)
                total  = cel_total.get(eid, 0)
                # Force continuation if mid-segment (consec==1, hasn't finished min 2-hour block)
                in_segment  = 1 if consec == 1 else 0
                # Then prefer those under the 2-consecutive cap (or solo manager)
                under_limit = 1 if (consec < 2 or len(available_mgrs) == 1) else 0
                # Fewest total CEL for even spread; latest end_time as tiebreaker
                return (in_segment, under_limit, -total, -(m.end_time.hour * 60 + m.end_time.minute))

            cel_pick = max(available_mgrs, key=_cel_priority)
            cel_eid  = cel_pick.employee_id
            hour_assignments[h][cel_eid] = 'CEL'
            cel_consecutive[cel_eid] = cel_consecutive.get(cel_eid, 0) + 1
            cel_total[cel_eid]       = cel_total.get(cel_eid, 0) + 1

            remaining_mgrs = [m for m in available_mgrs if m.employee_id != cel_eid]

            # All non-CEL managers get FLEX
            for mgr in remaining_mgrs:
                eid = mgr.employee_id
                hour_assignments[h][eid] = 'FLEX'
                cel_consecutive[eid] = 0

        rows = []
        for shift in shifts:
            sh, sm = shift.start_time.hour, shift.start_time.minute
            eh, em = shift.end_time.hour, shift.end_time.minute
            shift_label = f"{_fmt(sh, sm)}-{_fmt(eh, em)}"

            start_min = sh * 60 + sm
            end_min   = eh * 60 + em

            zones = {}
            for h in WORKBOOK_HOURS:
                if start_min < (h + 1) * 60 and end_min > h * 60:
                    if shift.role == 'BOH':
                        zones[str(h)] = 'BOH'
                    else:
                        zones[str(h)] = hour_assignments.get(h, {}).get(shift.employee_id, shift.role.upper())

            # Name stored as "Last, First" — show first name; add last initial if duplicate
            raw = shift.employee.name
            first, last_initial = _name_parts(raw)
            display = f"{first} {last_initial}" if first_name_counts[first] > 1 and last_initial else first

            rows.append({
                'name': display,
                'full_name': raw,
                'shift': shift_label,
                'role': shift.role,
                'zones': zones,
            })

        return Response({
            'day': day,
            'date': str(target_date),
            'hours': WORKBOOK_HOURS,
            'col_headers': WORKBOOK_COL_HEADERS,
            'rows': rows,
        })


class StaffZoneView(APIView):
    """GET /api/schedule/staff/ — list all employees with zone levels.
    PATCH /api/schedule/staff/<employee_id>/ — update zone fields."""
    permission_classes = [IsAuthenticated]

    def get(self, request, employee_id=None):
        # Auto-create StaffZone rows for any Employee that lacks one
        employees = Employee.objects.all().order_by('name')
        for emp in employees:
            StaffZone.objects.get_or_create(employee=emp)
        zones = StaffZone.objects.select_related('employee').order_by('employee__name')
        return Response(StaffZoneSerializer(zones, many=True).data)

    def patch(self, request, employee_id):
        try:
            zone = StaffZone.objects.get(employee_id=employee_id)
        except StaffZone.DoesNotExist:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        data = {k: v for k, v in request.data.items() if k in ZONE_FIELDS}
        for field, value in data.items():
            setattr(zone, field, int(value))
        zone.save(update_fields=list(data.keys()))
        return Response(StaffZoneSerializer(zone).data)


class InspectScheduleView(APIView):
    """POST an XLSX — returns raw sheet structure and first parsed records for debugging."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        xlsx_file = request.FILES.get('file')
        if not xlsx_file:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        tmp_path = _save_temp(xlsx_file)
        result = {}

        try:
            # Raw sheet peek
            wb = openpyxl.load_workbook(tmp_path, data_only=False)
            sheets = []
            for ws in wb.worksheets:
                rows_preview = []
                for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
                    rows_preview.append([repr(c) if c is not None else None for c in row])
                sheets.append({'name': ws.title, 'max_row': ws.max_row, 'max_col': ws.max_column, 'preview': rows_preview})
            result['sheets'] = sheets

            # Trace parse logic
            from .parsers import (
                _is_empty, _is_section_header, _is_column_header,
                _extract_date_cols, _str_cell, TIME_RANGE_RE, _cell,
            )
            wb2 = openpyxl.load_workbook(tmp_path, data_only=True)
            trace = []
            for ws in wb2.worksheets:
                raw_rows = [tuple(ws[rn]) for rn in range(1, ws.max_row + 1)]
                day_dates = {}
                day_cols  = {}
                for ri, row in enumerate(raw_rows):
                    row_label = ri + 1
                    if _is_empty(row):
                        trace.append({'row': row_label, 'type': 'empty'})
                        continue
                    if _is_section_header(row):
                        trace.append({'row': row_label, 'type': 'section_header', 'val': _str_cell(row, 0)})
                        continue
                    if _is_column_header(row):
                        date_row = raw_rows[ri + 1] if ri + 1 < len(raw_rows) else None
                        if date_row is not None:
                            day_dates, day_cols = _extract_date_cols(date_row)
                        date_row_raw = [repr(_cell(date_row, ci)) for ci in range(len(date_row))] if date_row is not None else []
                        trace.append({
                            'row': row_label, 'type': 'col_header',
                            'day_dates': {k: str(v) for k, v in day_dates.items()},
                            'day_cols': day_cols,
                            'date_row_raw': date_row_raw,
                        })
                        continue
                    name = _str_cell(row, 0)
                    if name:
                        time_cells = {}
                        for day, tc in day_cols.items():
                            raw = _str_cell(row, tc)
                            m = TIME_RANGE_RE.search(raw) if raw else None
                            time_cells[day] = {
                                'col': tc,
                                'raw': repr(raw),
                                'matched': bool(m),
                                'groups': list(m.groups()) if m else None,
                            }
                        trace.append({'row': row_label, 'type': 'employee', 'name': name, 'job': _str_cell(row, 3), 'time_cells': time_cells})
                    else:
                        trace.append({'row': row_label, 'type': 'continuation', 'col0': repr(_str_cell(row, 0))})
            result['trace'] = trace[:40]  # limit

            # Parser output
            records = parse_schedule_xlsx(tmp_path)
            result['parsed_count'] = len(records)
            result['sample_records'] = [
                {k: str(v) for k, v in r.items()} for r in records[:10]
            ]
        except Exception as exc:
            import traceback
            result['error'] = str(exc)
            result['traceback'] = traceback.format_exc()
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

        return Response(result)
