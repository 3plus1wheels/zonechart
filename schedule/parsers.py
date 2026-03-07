"""
schedule/parsers.py
--------------------
Parses the "Employee Schedule – Weekly" XLSX report produced by the WFM system.

Sheet layout (0-indexed columns):
  0  – Employee name (only on the employee's FIRST row in their block)
  3  – Primary Job   (same row as name)

  Day columns are AUTO-DETECTED from the date-header row each time a new
  "Employee / Primary Job" column-header is encountered.  The date values
  (e.g. "2/16/2026") appear directly above their respective time columns,
  so _extract_date_cols maps each detected date to the column index where
  it was found.  This is 100% layout-agnostic.

  Set PARSER_DEBUG=1 env var to print column-scan details to stderr.
"""

import re
import os
import sys
from datetime import datetime, date, time, timedelta
from typing import Optional

import openpyxl

PARSER_DEBUG = os.environ.get("PARSER_DEBUG", "0") == "1"

DAY_ORDER = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# ---------------------------------------------------------------------------
# BOH shift fingerprints  (start_time, end_time) — all times are naive
# ---------------------------------------------------------------------------
BOH_TIMES: set[tuple] = {
    (time(16, 30), time(21, 15)),   # 4:30 PM – 9:15 PM
    (time(8,   0), time(12, 45)),   # 8:00 AM – 12:45 PM
    (time(9,   0), time(13, 45)),   # 9:00 AM – 1:45 PM
}

# 10:00 AM – 2:45 PM is BOH on Sunday only; Stylist every other day
SUN_ONLY_BOH_TIMES: set[tuple] = {
    (time(10,  0), time(14, 45)),
}


def _normalize_role(raw_role: str, start_t: time, end_t: time, day_label: str = '', primary_job: str = '') -> str:
    """Map a raw WFM role string + shift times to one of: Stylist | CEL | BOH."""
    r = raw_role.lower()
    p = primary_job.lower()
    CEL_KEYWORDS = ("cel", "supervisor", "management")
    if any(kw in r for kw in CEL_KEYWORDS) or any(kw in p for kw in CEL_KEYWORDS):
        return "CEL"
    if "shipment" in r or "shipment" in p:
        return "BOH"
    if (start_t, end_t) in BOH_TIMES:
        return "BOH"
    if (start_t, end_t) in SUN_ONLY_BOH_TIMES and day_label == 'Sun':
        return "BOH"
    return "Stylist"


TIME_RANGE_RE = re.compile(
    r"(\d{1,2}:\d{2}\s*[AP]M)\s*[-\u2013\u2014]\s*(\d{1,2}:\d{2}\s*[AP]M)", re.IGNORECASE
)

SKIP_NAME_VALUES = {
    "Time Period :", "Time Period:", "Query :", "Query:",
    "Currency Code :", "Currency Code:",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_time(raw: str) -> Optional[time]:
    raw = raw.strip().replace('\u00a0', ' ').replace('\u2013', '-').replace('\u2014', '-').upper()
    for fmt in ("%I:%M %p", "%I:%M%p"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _parse_date_cell(v) -> Optional[date]:
    """Return a date from a cell value (datetime, date, serial float, or date string)."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # Excel serial date number (days since 1899-12-30)
    if isinstance(v, (int, float)) and 40000 < v < 60000:
        try:
            from datetime import timedelta
            return (date(1899, 12, 30) + timedelta(days=int(v)))
        except Exception:
            pass
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def _cell(row: tuple, idx: int):
    try:
        v = row[idx]
        return v.value if hasattr(v, "value") else v
    except IndexError:
        return None


def _str_cell(row: tuple, idx: int) -> str:
    v = _cell(row, idx)
    if v is None:
        return ""
    s = str(v).strip()
    s = s.replace('\u00a0', ' ').replace('\u200b', '').replace('\ufeff', '')
    return s


def _is_section_header(row: tuple) -> bool:
    first = _str_cell(row, 0)
    return first.startswith("LS&Co.") or first in SKIP_NAME_VALUES


def _is_column_header(row: tuple) -> bool:
    return _str_cell(row, 0).lower() == "employee"


def _extract_date_cols(row: tuple) -> tuple[dict[str, date], dict[str, int]]:
    """
    Scan ALL columns for date values. Returns two dicts keyed by day label:
      day_dates : { "Mon": date(…), … }
      day_cols  : { "Mon": 9, "Tue": 11, … }  ← actual column indices
    Dates are assigned left-to-right to DAY_ORDER, so the column a date
    lands in becomes that day's time column.
    """
    found: list[tuple[int, date]] = []  # (col_index, date)
    seen: set[date] = set()
    for i in range(len(row)):
        d = _parse_date_cell(_cell(row, i))
        if d is not None and d not in seen:
            found.append((i, d))
            seen.add(d)

    day_dates: dict[str, date] = {}
    day_cols:  dict[str, int]  = {}
    for idx, (col, d) in enumerate(found):
        if idx >= len(DAY_ORDER):
            break
        label = DAY_ORDER[idx]
        day_dates[label] = d
        day_cols[label]  = col

    if PARSER_DEBUG:
        print("[PARSER] _extract_date_cols:", file=sys.stderr)
        for label in DAY_ORDER:
            if label in day_cols:
                print(f"  {label}: col={day_cols[label]:>2}  date={day_dates[label]}", file=sys.stderr)
            else:
                print(f"  {label}: NOT FOUND", file=sys.stderr)

    return day_dates, day_cols


def _is_empty(row: tuple) -> bool:
    return all(_cell(row, i) is None for i in range(len(row)))


# ---------------------------------------------------------------------------
# Employee block extractor
# ---------------------------------------------------------------------------

def _extract_employee_shifts(
    block: list[tuple],
    day_dates: dict[str, date],
    day_cols: dict[str, int],
    employee_name: str,
    primary_job: str,
) -> list[dict]:
    shifts = []
    pairs = [(block[i], block[i + 1] if i + 1 < len(block) else None)
             for i in range(0, len(block), 2)]

    if PARSER_DEBUG:
        print(f"[PARSER] Employee: {employee_name!r}  block_rows={len(block)}", file=sys.stderr)

    for pair_i, (time_row, role_row) in enumerate(pairs):
        if PARSER_DEBUG:
            print(f"  pair {pair_i}:", file=sys.stderr)
            for label in DAY_ORDER:
                col = day_cols.get(label)
                if col is not None:
                    print(f"    {label} col={col:>2}: {_str_cell(time_row, col)!r}", file=sys.stderr)

        for day_label, time_col in day_cols.items():
            raw_time = _str_cell(time_row, time_col)
            m = TIME_RANGE_RE.search(raw_time)
            if not m:
                continue

            start_t = _parse_time(m.group(1))
            end_t = _parse_time(m.group(2))
            if not start_t or not end_t:
                continue

            shift_date = day_dates.get(day_label)
            if not shift_date:
                continue

            raw_role = _str_cell(role_row, time_col) if role_row else ""
            role = _normalize_role(raw_role, start_t, end_t, day_label, primary_job)

            shifts.append({
                "employee_name": employee_name,
                "primary_job": primary_job,
                "day_label": day_label,
                "date": shift_date,
                "start_time": start_t,
                "end_time": end_t,
                "role": role,
            })

    return shifts


# ---------------------------------------------------------------------------
# Shift merger
# ---------------------------------------------------------------------------

def _merge_close_shifts(shifts: list[dict], gap_minutes: int = 30) -> list[dict]:
    """
    For each (employee, date) group, sort shifts by start_time and merge any
    consecutive pair whose gap is <= gap_minutes minutes into a single shift.
    - end_time  : taken from the later shift
    - role      : combined with ' / ' if the two roles differ
    - is_closing: True if either shift was closing
    """
    from collections import defaultdict
    max_gap = timedelta(minutes=gap_minutes)

    # Group preserving original relative order, then sort within group
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for s in shifts:
        groups[(s["employee_name"], s["date"])].append(s)

    merged_all: list[dict] = []
    for group in groups.values():
        group.sort(key=lambda s: s["start_time"])
        merged: list[dict] = []
        for s in group:
            if merged and (
                datetime.combine(s["date"], s["start_time"]) -
                datetime.combine(s["date"], merged[-1]["end_time"])
                <= max_gap
            ):
                prev = merged[-1]
                # Extend end time
                if s["end_time"] > prev["end_time"]:
                    prev["end_time"] = s["end_time"]
                # Combine role labels if different
                if s["role"] and s["role"] != prev["role"]:
                    prev["role"] = " / ".join(filter(None, [prev["role"], s["role"]]))
            else:
                merged.append(dict(s))  # copy so original is untouched
        merged_all.extend(merged)

    return merged_all


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_schedule_xlsx(filepath: str) -> list[dict]:
    """Parse every sheet in the XLSX and return a flat list of shift dicts."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_shifts: list[dict] = []

    for ws in wb.worksheets:
        rows = [tuple(ws[row_num]) for row_num in range(1, ws.max_row + 1)]
        all_shifts.extend(_parse_rows(rows))

    return _merge_close_shifts(all_shifts)


def _parse_rows(rows: list[tuple]) -> list[dict]:
    """
    Walk the row list handling multiple sub-table blocks within a single sheet.
    A sheet may contain several "Employee / dates / employee rows..." blocks
    (one per store location, etc.).  Each block is re-started when we see a
    fresh column-header row.
    """
    shifts: list[dict] = []
    day_dates: dict[str, date] = {}
    day_cols:  dict[str, int]  = {}
    i = 0

    while i < len(rows):
        row = rows[i]

        if _is_empty(row):
            i += 1
            continue

        # Must check column header BEFORE section header
        if _is_column_header(row):
            # The very next row holds the date values for this block
            if i + 1 < len(rows):
                day_dates, day_cols = _extract_date_cols(rows[i + 1])
            i += 2
            continue

        if _is_section_header(row):
            i += 1
            continue

        name = _str_cell(row, 0)
        if name and name not in SKIP_NAME_VALUES:
            primary_job = _str_cell(row, 3)

            # Collect the continuation rows belonging to this employee
            block = [row]
            j = i + 1
            while j < len(rows):
                nxt = rows[j]
                if (_is_empty(nxt) or _is_section_header(nxt)
                        or _is_column_header(nxt) or _str_cell(nxt, 0)):
                    break
                block.append(nxt)
                j += 1

            shifts.extend(
                _extract_employee_shifts(block, day_dates, day_cols, name, primary_job)
            )
            i = j
            continue

        i += 1

    return shifts
